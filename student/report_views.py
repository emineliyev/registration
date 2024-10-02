from datetime import date
import openpyxl
from django.db.models import Q, OuterRef, Exists
from django.db.models import Sum
from django.http import HttpResponse
from django.utils import timezone
from django.views.generic import View
from student.models import Student, TuitionFeeSplit
from student.views import BaseStudentListView


class LatePaymentListView(BaseStudentListView):
    template_name = 'student/report/late_payments.html'

    # Метод для обновления статуса должников
    def update_debtor_statuses(self):
        today = timezone.localdate()
        overdue_payments_subquery = TuitionFeeSplit.objects.filter(
            student_id=OuterRef('pk'),
            payment_date__isnull=True,
            installment_date__lt=today
        )
        Student.objects.update(is_debtor=Exists(overdue_payments_subquery))

    def get_filters(self):
        start_date = self.request.GET.get('overdue_date__gte')
        end_date = self.request.GET.get('overdue_date__lte')

        filters = Q()

        if start_date and end_date:
            overdue_students = TuitionFeeSplit.objects.filter(
                payment_date__isnull=True,
                installment_date__lt=date.today(),
                installment_date__gte=start_date,
                installment_date__lte=end_date
            ).values_list('student_id', flat=True)
            filters &= Q(id__in=overdue_students)

        return filters

    def get_queryset(self):
        self.update_debtor_statuses()
        filters = self.get_filters()
        queryset = super().get_queryset().filter(is_debtor=True).filter(filters).distinct()
        return queryset

    def get_total_amounts(self, students):
        total_paid_sum = sum(getattr(student, 'total_paid', 0) for student in students)
        total_debt_amount_sum = sum(getattr(student, 'debt_amount', 0) for student in students)
        return total_paid_sum, total_debt_amount_sum

    def get_student_payment_data(self, student_ids):
        tuition_fee_splits = TuitionFeeSplit.objects.filter(student_id__in=student_ids).values(
            'student_id', 'installment_date', 'installment_amount', 'payment_date', 'payment_amount'
        )

        student_payment_data = {}
        for split in tuition_fee_splits:
            student_id = split['student_id']
            if student_id not in student_payment_data:
                student_payment_data[student_id] = {
                    'total_paid': 0,
                    'debt_amount': 0,
                    'debt_months': 0,
                    'overdue_installment_dates': []
                }
            if split['payment_date']:
                student_payment_data[student_id]['total_paid'] += split['payment_amount']
            else:
                if split['installment_date'] < date.today():
                    student_payment_data[student_id]['debt_months'] += 1
                    student_payment_data[student_id]['debt_amount'] += split['installment_amount']
                    student_payment_data[student_id]['overdue_installment_dates'].append(split['installment_date'])

        return student_payment_data

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        students = self.get_queryset()
        student_ids = list(students.values_list('id', flat=True))

        student_payment_data = self.get_student_payment_data(student_ids)

        for student in students:
            student_id = student.id
            payment_data = student_payment_data.get(student_id, {})
            total_paid = payment_data.get('total_paid', 0)
            debt_amount = payment_data.get('debt_amount', 0)
            debt_months = payment_data.get('debt_months', 0)
            overdue_installment_dates = ", ".join(
                map(str, payment_data.get('overdue_installment_dates', [])))

            student.total_paid = total_paid
            student.debt_amount = debt_amount
            student.debt_months = debt_months
            student.overdue_installment_dates = overdue_installment_dates

        total_paid_sum, total_debt_amount_sum = self.get_total_amounts(students)

        totals = students.aggregate(total_annual_fee=Sum('annual_fee'))

        context['students'] = students
        context['total_paid_sum'] = total_paid_sum
        context['total_debt_amount_sum'] = total_debt_amount_sum
        context['total_annual_fee_sum'] = totals['total_annual_fee'] or 0

        return context


class DailyPayments(BaseStudentListView):
    template_name = 'student/report/daily_payments.html'
    paginate_by = 25

    def get_filters(self):
        payment_date_gte = self.request.GET.get('payment_date_gte')
        payment_date_lte = self.request.GET.get('payment_date_lte')

        filters = Q()

        if payment_date_gte and payment_date_lte:
            filters &= Q(tuitionfeesplit__payment_date__gte=payment_date_gte) & Q(
                tuitionfeesplit__payment_date__lte=payment_date_lte)
        elif payment_date_gte:
            filters &= Q(tuitionfeesplit__payment_date__gte=payment_date_gte)
        elif payment_date_lte:
            filters &= Q(tuitionfeesplit__payment_date__lte=payment_date_lte)
        else:
            payment_date = self.request.GET.get('payment_date', date.today().strftime('%Y-%m-%d'))
            filters &= Q(tuitionfeesplit__payment_date=payment_date)

        year = self.request.GET.get('year')
        work_number = self.request.GET.get('work_number__icontains')
        first_name = self.request.GET.get('first_name__icontains')
        last_name = self.request.GET.get('last_name__icontains')
        student_class = self.request.GET.getlist('student_class')  # Изменено на getlist
        group = self.request.GET.getlist('group')  # Изменено на getlist
        overdue_date_gte = self.request.GET.get('overdue_date__gte')
        overdue_date_lte = self.request.GET.get('overdue_date__lte')

        if year:
            filters &= Q(year_id=year)
        if work_number:
            filters &= Q(work_number__icontains=work_number)
        if first_name:
            filters &= Q(first_name__icontains=first_name)
        if last_name:
            filters &= Q(last_name__icontains=last_name)
        if student_class:
            filters &= Q(student_class_id__in=student_class)  # Изменено на __in
        if group:
            filters &= Q(group_id__in=group)  # Изменено на __in
        if overdue_date_gte:
            filters &= Q(tuitionfeesplit__installment_date__gte=overdue_date_gte)
        if overdue_date_lte:
            filters &= Q(tuitionfeesplit__installment_date__lte=overdue_date_lte)

        return filters

    def get_queryset(self):
        filters = self.get_filters()
        queryset = super().get_queryset().select_related('student_class').prefetch_related('tuitionfeesplit_set')
        queryset = queryset.filter(filters).distinct().annotate(total_payment=Sum('tuitionfeesplit__payment_amount'))
        return queryset

    def get_total_amount(self):
        filters = self.get_filters()
        queryset = super().get_queryset().select_related('student_class').prefetch_related('tuitionfeesplit_set')
        total_amount = queryset.filter(filters).aggregate(total=Sum('tuitionfeesplit__payment_amount'))['total'] or 0
        return total_amount

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        students = context.get('students', self.get_queryset())
        payment_date_gte = self.request.GET.get('payment_date_gte')
        payment_date_lte = self.request.GET.get('payment_date_lte')

        student_ids = students.values_list('id', flat=True)
        payments = TuitionFeeSplit.objects.filter(student_id__in=student_ids)

        if payment_date_gte and payment_date_lte:
            payments = payments.filter(payment_date__gte=payment_date_gte, payment_date__lte=payment_date_lte)
        elif payment_date_gte:
            payments = payments.filter(payment_date__gte=payment_date_gte)
        elif payment_date_lte:
            payments = payments.filter(payment_date__lte=payment_date_lte)
        else:
            payment_date = self.request.GET.get('payment_date', date.today().strftime('%Y-%m-%d'))
            payments = payments.filter(payment_date=payment_date)

        payments_dict = {}
        for payment in payments:
            if payment.student_id not in payments_dict:
                payments_dict[payment.student_id] = []
            payments_dict[payment.student_id].append(payment)

        for student in students:
            student.payments = payments_dict.get(student.id, [])

        total_amount = self.get_total_amount()

        context['students'] = students
        context['payment_date_gte'] = payment_date_gte
        context['payment_date_lte'] = payment_date_lte
        context['total_amount'] = total_amount
        return context


class StudentReportListView(BaseStudentListView):
    template_name = 'student/report/student_report.html'
    paginate_by = 25

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        students = context['object_list']  # Используем пагинированный список

        all_students = self.get_queryset()
        totals = all_students.aggregate(total_annual_fee=Sum('annual_fee'))
        context['annual_fee_total'] = totals['total_annual_fee']

        student_ids = list(all_students.values_list('id', flat=True))
        tuition_fee_splits = TuitionFeeSplit.objects.filter(student_id__in=student_ids).values(
            'student_id', 'payment_date', 'payment_amount', 'installment_amount'
        )

        student_payment_data = {}
        for split in tuition_fee_splits:
            student_id = split['student_id']
            if student_id not in student_payment_data:
                student_payment_data[student_id] = {
                    'total_paid': 0,
                    'debt_amount': 0,
                }
            if split['payment_date']:
                student_payment_data[student_id]['total_paid'] += split['payment_amount']
            else:
                student_payment_data[student_id]['debt_amount'] += split['installment_amount']

        for student in all_students:
            student_id = student.id
            total_paid = student_payment_data.get(student_id, {}).get('total_paid', 0)
            debt_amount = student_payment_data.get(student_id, {}).get('debt_amount', 0)

            student.total_paid = total_paid
            student.debt_amount = debt_amount

        for student in students:
            student_id = student.id
            student.total_paid = student_payment_data.get(student_id, {}).get('total_paid', 0)
            student.debt_amount = student_payment_data.get(student_id, {}).get('debt_amount', 0)

        context['total_paid_sum'] = sum(student.total_paid for student in all_students)
        context['total_debt_sum'] = sum(student.debt_amount for student in all_students)

        return context


class ExportLatePaymentsExcelView(View):
    def get(self, request, *args, **kwargs):
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Late Payments"

        columns = [
            "İş nömrə", "Ad", "Soyad", "Sinif", "Qrup", "Sinif коду", "Bölmə",
            "Məhsul şəxsin telefonu", "Ümumi məbləğ", "Ödənən məbləğ", "Gecikən məbləğ", "Gecikən taksit sayı",
            "Gecikən taksit tarixləri"
        ]
        row_num = 1

        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title

        queryset = Student.objects.filter(is_debtor=True).select_related('student_class', 'group', 'class_number')

        year = request.GET.get('year')
        work_number = request.GET.get('work_number__icontains')
        first_name = request.GET.get('first_name__icontains')
        last_name = request.GET.get('last_name__icontains')
        student_class = request.GET.getlist('student_class')  # Изменено на getlist
        group = request.GET.getlist('group')  # Изменено на getlist
        class_number = request.GET.get('class_number')
        section = request.GET.get('section')
        overdue_date_gte = request.GET.get('overdue_date__gte')
        overdue_date_lte = request.GET.get('overdue_date__lte')

        filters = Q()

        if year:
            filters &= Q(year_id=year)
        if work_number:
            filters &= Q(work_number__icontains=work_number)
        if first_name:
            filters &= Q(first_name__icontains=first_name)
        if last_name:
            filters &= Q(last_name__icontains=last_name)
        if student_class:
            filters &= Q(student_class_id__in=student_class)  # Изменено на __in
        if group:
            filters &= Q(group_id__in=group)  # Изменено на __in
        if class_number:
            filters &= Q(class_number_id=class_number)
        if section:
            filters &= Q(section=section)
        if overdue_date_gte and overdue_date_lte:
            overdue_students = TuitionFeeSplit.objects.filter(
                payment_date__isnull=True,
                installment_date__lt=date.today(),
                installment_date__gte=overdue_date_gte,
                installment_date__lte=overdue_date_lte
            ).values_list('student_id', flat=True)
            filters &= Q(id__in=overdue_students)

        queryset = queryset.filter(filters)

        student_ids = list(queryset.values_list('id', flat=True))

        tuition_fee_splits = TuitionFeeSplit.objects.filter(student_id__in=student_ids).values(
            'student_id', 'installment_date', 'installment_amount', 'payment_date', 'payment_amount'
        )

        student_payment_data = {}
        for split in tuition_fee_splits:
            student_id = split['student_id']
            if student_id not in student_payment_data:
                student_payment_data[student_id] = {
                    'total_paid': 0,
                    'unpaid_installments': [],
                    'debt_months': 0,
                    'debt_amount': 0,
                    'overdue_installment_dates': []
                }
            if split['payment_date']:
                student_payment_data[student_id]['total_paid'] += split['payment_amount']
            else:
                if split['installment_date'] < date.today():
                    student_payment_data[student_id]['debt_months'] += 1
                    student_payment_data[student_id]['debt_amount'] += split['installment_amount']
                    student_payment_data[student_id]['overdue_installment_dates'].append(split['installment_date'])

        totals = queryset.aggregate(total_annual_fee=Sum('annual_fee'))

        for student in queryset:
            student_id = student.id
            total_paid = student_payment_data[student_id]['total_paid']
            debt_months = student_payment_data[student_id]['debt_months']
            debt_amount = student_payment_data[student_id]['debt_amount']
            overdue_installment_dates = ", ".join(
                map(str, student_payment_data[student_id]['overdue_installment_dates']))
            annual_fee = student.annual_fee or 0

            row_num += 1
            row = [
                student.work_number, student.first_name, student.last_name,
                student.student_class.name, student.group.name, student.class_number.number,
                student.get_section_display(), student.responsible_phone_1,
                annual_fee, total_paid, debt_amount, debt_months,
                overdue_installment_dates
            ]

            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

        row_num += 1
        total_row = [
            "Cəmi", "", "", "", "", "", "", "", totals['total_annual_fee'] or 0, "", "", "", ""
        ]
        for col_num, cell_value in enumerate(total_row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=geciken_odenisler.xlsx'
        workbook.save(response)
        return response


class ExportDailyPaymentsExcelView(View):
    def get(self, request, *args, **kwargs):
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Daily Payments"

        columns = ["İş nömrə", "Ad", "Soyad", "Sinif", "Tarix", "Məbləğ"]
        row_num = 1

        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title

        daily_payments_view = DailyPayments()
        daily_payments_view.request = request
        filters = daily_payments_view.get_filters()

        queryset = Student.objects.filter(filters).distinct().select_related('student_class').prefetch_related(
            'tuitionfeesplit_set')

        student_ids = queryset.values_list('id', flat=True)
        payments = TuitionFeeSplit.objects.filter(student_id__in=student_ids)

        payment_date_gte = request.GET.get('payment_date_gte')
        payment_date_lte = request.GET.get('payment_date_lte')

        if payment_date_gte and payment_date_lte:
            payments = payments.filter(payment_date__gte=payment_date_gte, payment_date__lte=payment_date_lte)
        elif payment_date_gte:
            payments = payments.filter(payment_date__gte=payment_date_gte)
        elif payment_date_lte:
            payments = payments.filter(payment_date__lte=payment_date_lte)
        else:
            payment_date = request.GET.get('payment_date', date.today().strftime('%Y-%m-%d'))
            payments = payments.filter(payment_date=payment_date)

        total_amount = 0
        payments_dict = {}
        for payment in payments:
            if payment.student_id not in payments_dict:
                payments_dict[payment.student_id] = []
            payments_dict[payment.student_id].append(payment)

        for student in queryset:
            student.payments = payments_dict.get(student.id, [])
            for payment in student.payments:
                row_num += 1
                row = [student.work_number, student.first_name, student.last_name, student.student_class.name,
                       payment.payment_date.strftime('%Y-%m-%d'), payment.payment_amount]
                for col_num, cell_value in enumerate(row, 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.value = cell_value
                total_amount += payment.payment_amount

        row_num += 1
        worksheet.cell(row=row_num, column=5).value = "Cəmi"
        worksheet.cell(row=row_num, column=6).value = total_amount

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response[
            'Content-Disposition'] = f'attachment; filename=gundelik_odenisler_{date.today().strftime("%Y-%m-%d")}.xlsx'
        workbook.save(response)
        return response


class ExportStudentReportExcelView(View):
    def get(self, request, *args, **kwargs):
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Student Report"

        columns = [
            "İş nömrə", "Ad", "Soyad", "Sinif", "Qeydiyyat statusu", "İllik məbləğ", "Ödənilən məbləğ", "Qalıq"
        ]
        row_num = 1

        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title

        queryset = Student.objects.all()
        filters = self.request.GET

        if filters:
            if filters.get('year'):
                queryset = queryset.filter(year_id=filters.get('year'))
            if filters.get('work_number__icontains'):
                queryset = queryset.filter(work_number__icontains=filters.get('work_number__icontains'))
            if filters.get('first_name__icontains'):
                queryset = queryset.filter(first_name__icontains=filters.get('first_name__icontains'))
            if filters.get('last_name__icontains'):
                queryset = queryset.filter(last_name__icontains=filters.get('last_name__icontains'))
            if filters.get('student_class'):
                queryset = queryset.filter(student_class_id__in=filters.getlist('student_class'))  # Изменено на __in
            if filters.get('registration_status'):
                queryset = queryset.filter(registration_status=filters.get('registration_status'))
            if filters.get('created_at__gte'):
                queryset = queryset.filter(created_at__gte=filters.get('created_at__gte'))
            if filters.get('created_at__lte'):
                queryset = queryset.filter(created_at__lte=filters.get('created_at__lte'))

        student_ids = list(queryset.values_list('id', flat=True))

        tuition_fee_splits = TuitionFeeSplit.objects.filter(student_id__in=student_ids).values(
            'student_id', 'installment_date', 'installment_amount', 'payment_date', 'payment_amount'
        )

        student_payment_data = {}
        for split in tuition_fee_splits:
            student_id = split['student_id']
            if student_id not in student_payment_data:
                student_payment_data[student_id] = {
                    'total_paid': 0,
                    'debt_amount': 0,
                }
            if split['payment_date']:
                student_payment_data[student_id]['total_paid'] += split['payment_amount']
            else:
                student_payment_data[student_id]['debt_amount'] += split['installment_amount']

        for student in queryset:
            student_id = student.id
            total_paid = student_payment_data.get(student_id, {}).get('total_paid', 0)
            debt_amount = student_payment_data.get(student_id, {}).get('debt_amount', 0)
            annual_fee = student.annual_fee or 0

            row_num += 1
            row = [
                student.work_number, student.first_name, student.last_name,
                student.student_class.name, student.get_registration_status_display(), annual_fee, total_paid,
                debt_amount
            ]

            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=sagird_uzre_hesabat.xlsx'
        workbook.save(response)
        return response
